import openpyxl
path = "C:/Users/Dell/Desktop/New folder/login.xlsx"    # providing the path
workbook = openpyxl.load_workbook(path)                 # opening the excel
sheet = workbook.active                                 # opening the sheet from excel
#sheet = workbook.get_sheet_by_name(Sheet1)
rows = sheet.max_row                                    # opening the rows from the excel
cols = sheet.max_column                                 # opening the columns from the excel
print(rows)
print(cols)

# reading the data from it
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="  ")
    print()
